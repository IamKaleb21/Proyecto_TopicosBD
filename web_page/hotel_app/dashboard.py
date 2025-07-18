import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from pymongo import MongoClient
from datetime import datetime, timedelta
from bson.objectid import ObjectId
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

# Configuración de la página
st.set_page_config(
    page_title="Dashboard Hotel Costa del Inka",
    page_icon="🏨",
    layout="wide"
)

# Conexión a MongoDB
@st.cache_resource
def init_connection():
    return MongoClient('mongodb://localhost:27017/')

client = init_connection()
db = client['CostaDelInkaDB']

# Funciones de consulta
@st.cache_data(ttl=600)
def get_canales_reserva():
    """Obtener lista única de canales de reserva."""
    canales = list(db.Reservas.distinct("canal_reserva"))
    return ["Todos"] + [c for c in canales if c]

@st.cache_data(ttl=600)
def get_tipos_habitacion():
    """Obtener lista única de tipos de habitación."""
    tipos = list(db.DetallesReserva.distinct("tipo_habitacion_reservada"))
    return ["Todos"] + [t for t in tipos if t]

def get_metricas_principales(fecha_inicio, fecha_fin, canal_reserva=None):
    """Obtener métricas principales."""
    filtro = {
        "fecha_llegada": {
            "$gte": fecha_inicio,
            "$lte": fecha_fin
        }
    }
    
    if canal_reserva and canal_reserva != "Todos":
        filtro["canal_reserva"] = canal_reserva
    
    total_reservas = db.Reservas.count_documents(filtro)
    total_clientes = db.Clientes.count_documents({})
    
    # Calcular tasa de cancelación
    reservas_canceladas = db.Reservas.count_documents({**filtro, "fue_cancelada": True})
    tasa_cancelacion = (reservas_canceladas / total_reservas * 100) if total_reservas > 0 else 0
    
    # Calcular ADR promedio
    pipeline_adr = [
        {"$match": filtro},
        {"$group": {"_id": None, "adr_promedio": {"$avg": "$adr"}}}
    ]
    resultado_adr = list(db.Reservas.aggregate(pipeline_adr))
    adr_promedio = resultado_adr[0]["adr_promedio"] if resultado_adr else 0
    
    return total_reservas, total_clientes, tasa_cancelacion, adr_promedio

def get_datos_reservas(fecha_inicio, fecha_fin, canal_reserva=None):
    """Obtener datos de reservas para gráficos."""
    match_stage = {
        "fecha_llegada": {
            "$gte": fecha_inicio,
            "$lte": fecha_fin
        }
    }
    
    if canal_reserva and canal_reserva != "Todos":
        match_stage["canal_reserva"] = canal_reserva
    
    # Reservas por día
    pipeline_dias = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$fecha_llegada"}},
                "total": {"$sum": 1}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    # Distribución por canal
    pipeline_canales = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": "$canal_reserva",
                "total": {"$sum": 1}
            }
        }
    ]
    
    # Tipos de habitación
    pipeline_habitaciones = [
        {
            "$lookup": {
                "from": "DetallesReserva",
                "localField": "detalle_reserva_id",
                "foreignField": "_id",
                "as": "detalle"
            }
        },
        {"$unwind": "$detalle"},
        {"$match": match_stage},
        {
            "$group": {
                "_id": "$detalle.tipo_habitacion_reservada",
                "total": {"$sum": 1}
            }
        }
    ]
    
    reservas_por_dia = list(db.Reservas.aggregate(pipeline_dias))
    distribucion_canales = list(db.Reservas.aggregate(pipeline_canales))
    tipos_habitacion = list(db.Reservas.aggregate(pipeline_habitaciones))
    
    return reservas_por_dia, distribucion_canales, tipos_habitacion

def get_datos_detallados(fecha_inicio, fecha_fin, canal_reserva=None):
    """Obtener datos detallados para análisis avanzado."""
    match_stage = {
        "fecha_llegada": {
            "$gte": fecha_inicio,
            "$lte": fecha_fin
        }
    }
    
    if canal_reserva and canal_reserva != "Todos":
        match_stage["canal_reserva"] = canal_reserva
    
    # Pipeline completo con joins
    pipeline = [
        {"$match": match_stage},
        {
            "$lookup": {
                "from": "DetallesReserva",
                "localField": "detalle_reserva_id",
                "foreignField": "_id",
                "as": "detalle"
            }
        },
        {"$unwind": "$detalle"},
        {
            "$lookup": {
                "from": "Clientes",
                "localField": "cliente_id",
                "foreignField": "_id",
                "as": "cliente"
            }
        },
        {"$unwind": "$cliente"},
        {
            "$project": {
                "fecha_llegada": 1,
                "fecha_salida": 1,
                "noches_estadia": 1,
                "adr": 1,
                "canal_reserva": 1,
                "fue_cancelada": 1,
                "estado_reserva": 1,
                "tiempo_anticipacion_reserva_dias": 1,
                "tipo_habitacion": "$detalle.tipo_habitacion_reservada",
                "tipo_cliente": "$detalle.tipo_cliente_en_reserva",
                "pais_origen": "$detalle.pais_origen_reserva",
                "es_huesped_recurrente": "$detalle.es_huesped_recurrente_al_reservar",
                "cambios_en_reserva": "$detalle.cambios_en_reserva",
                "nombre_cliente": "$cliente.nombre_completo",
                "email_cliente": "$cliente.email",
                "pais_cliente": "$cliente.pais_origen_cliente"
            }
        }
    ]
    
    return list(db.Reservas.aggregate(pipeline))

def get_analisis_clientes():
    """Obtener análisis de clientes."""
    # Clientes por país
    pipeline_paises = [
        {"$group": {
            "_id": "$pais_origen_cliente",
            "total": {"$sum": 1}
        }},
        {"$sort": {"total": -1}},
        {"$limit": 10}
    ]
    
    # Huéspedes recurrentes vs nuevos
    pipeline_recurrentes = [
        {"$group": {
            "_id": "$es_huesped_recurrente_historico",
            "total": {"$sum": 1}
        }}
    ]
    
    # Clientes con más reservas
    pipeline_top_clientes = [
        {"$project": {
            "nombre": "$nombre_completo",
            "total_reservas": {"$size": "$historial_ids_reservas"}
        }},
        {"$sort": {"total_reservas": -1}},
        {"$limit": 10}
    ]
    
    paises = list(db.Clientes.aggregate(pipeline_paises))
    recurrentes = list(db.Clientes.aggregate(pipeline_recurrentes))
    top_clientes = list(db.Clientes.aggregate(pipeline_top_clientes))
    
    return paises, recurrentes, top_clientes

def get_analisis_ocupacion(fecha_inicio, fecha_fin):
    """Obtener análisis de ocupación."""
    # Ocupación por mes
    pipeline_ocupacion = [
        {"$match": {
            "fecha_llegada": {
                "$gte": fecha_inicio,
                "$lte": fecha_fin
            }
        }},
        {
            "$group": {
                "_id": {
                    "anio": {"$year": "$fecha_llegada"},
                    "mes": {"$month": "$fecha_llegada"}
                },
                "total_reservas": {"$sum": 1},
                "total_noches": {"$sum": "$noches_estadia"},
                "adr_promedio": {"$avg": "$adr"}
            }
        },
        {"$sort": {"_id.anio": 1, "_id.mes": 1}}
    ]
    
    return list(db.Reservas.aggregate(pipeline_ocupacion))

def clean_data_for_excel(df):
    """Limpiar datos para exportación a Excel."""
    df_clean = df.copy()
    
    for column in df_clean.columns:
        # Convertir ObjectId a string
        if df_clean[column].dtype == 'object':
            df_clean[column] = df_clean[column].astype(str)
        
        # Convertir fechas a string si es necesario
        if 'fecha' in column.lower() and df_clean[column].dtype == 'datetime64[ns]':
            df_clean[column] = df_clean[column].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convertir booleanos a string
        if df_clean[column].dtype == 'bool':
            df_clean[column] = df_clean[column].map({True: 'Sí', False: 'No'})
    
    return df_clean

# Título principal
st.title("🏨 Dashboard Hotel Costa del Inka")
st.markdown("Sistema de Análisis de Reservas y Ocupación")

# Filtros en la barra lateral
with st.sidebar:
    st.header("Filtros")
    
    # Obtener fechas disponibles en la BD
    pipeline_fechas = [
        {"$group": {
            "_id": None,
            "fecha_min": {"$min": "$fecha_llegada"},
            "fecha_max": {"$max": "$fecha_llegada"}
        }}
    ]
    fechas_info = list(db.Reservas.aggregate(pipeline_fechas))
    
    if fechas_info:
        fecha_min_bd = fechas_info[0]["fecha_min"].date()
        fecha_max_bd = fechas_info[0]["fecha_max"].date()
        
        # Selector de fechas
        fecha_inicio = st.date_input(
            "Fecha Inicio",
            value=fecha_min_bd,
            min_value=fecha_min_bd,
            max_value=fecha_max_bd
        )
        
        fecha_fin = st.date_input(
            "Fecha Fin",
            value=fecha_max_bd,
            min_value=fecha_inicio,
            max_value=fecha_max_bd
        )
        
        # Convertir a datetime
        fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
        fecha_fin = datetime.combine(fecha_fin, datetime.max.time())
        
        # Selector de canal
        canal_reserva = st.selectbox(
            "Canal de Reserva",
            options=get_canales_reserva()
        )
    else:
        st.error("No se encontraron datos de fechas en la base de datos")
        st.stop()

# Crear pestañas
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📊 Resumen General", 
    "📈 Análisis Detallado", 
    "👥 Análisis de Clientes", 
    "🏠 Análisis de Ocupación",
    "➕ Nueva Reserva",
    "🔍 Búsqueda y Gestión",
    "📈 Reportes y Exportación"
])

# Pestaña 1: Resumen General
with tab1:
    st.header("📊 Resumen General")
    
    # Métricas principales
    total_reservas, total_clientes, tasa_cancelacion, adr_promedio = get_metricas_principales(
        fecha_inicio, fecha_fin, canal_reserva
    )
    
    # Mostrar métricas en columnas
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Reservas", f"{total_reservas:,}")
    
    with col2:
        st.metric("Total Clientes", f"{total_clientes:,}")
    
    with col3:
        st.metric("Tasa de Cancelación", f"{tasa_cancelacion:.1f}%")
    
    with col4:
        st.metric("ADR Promedio", f"${adr_promedio:.2f}")
    
    # --- Análisis de Cancelación de Reservas ---
    st.subheader("❌ Análisis de Cancelación de Reservas")
    # Gráfico de pastel: canceladas vs no canceladas
    pipeline_cancel = [
        {"$match": {
            "fecha_llegada": {"$gte": fecha_inicio, "$lte": fecha_fin},
            **({"canal_reserva": canal_reserva} if canal_reserva and canal_reserva != "Todos" else {})
        }},
        {"$group": {"_id": "$fue_cancelada", "total": {"$sum": 1}}}
    ]
    cancel_data = list(db.Reservas.aggregate(pipeline_cancel))
    if cancel_data:
        df_cancel = pd.DataFrame(cancel_data)
        df_cancel['_id'] = df_cancel['_id'].map({True: 'Canceladas', False: 'No Canceladas'})
        fig_cancel = px.pie(
            df_cancel,
            values="total",
            names="_id",
            title="Porcentaje de Reservas Canceladas vs No Canceladas",
            color_discrete_map={"Canceladas": "#EF553B", "No Canceladas": "#00CC96"}
        )
        st.plotly_chart(fig_cancel, use_container_width=True)
    else:
        st.warning("No hay datos de cancelación para el período seleccionado")

    # Gráfico de barras: número de cancelaciones por mes
    pipeline_cancel_rate = [
        {"$match": {
            "fecha_llegada": {"$gte": fecha_inicio, "$lte": fecha_fin},
            **({"canal_reserva": canal_reserva} if canal_reserva and canal_reserva != "Todos" else {})
        }},
        {"$group": {
            "_id": {"anio": {"$year": "$fecha_llegada"}, "mes": {"$month": "$fecha_llegada"}},
            "total": {"$sum": 1},
            "canceladas": {"$sum": {"$cond": ["$fue_cancelada", 1, 0]}}
        }},
        {"$sort": {"_id.anio": 1, "_id.mes": 1}}
    ]
    cancel_rate_data = list(db.Reservas.aggregate(pipeline_cancel_rate))
    if cancel_rate_data:
        df_cancel_rate = pd.DataFrame(cancel_rate_data)
        df_cancel_rate['fecha'] = pd.to_datetime(df_cancel_rate['_id'].apply(lambda x: f"{x['anio']}-{x['mes']:02d}-01"))
        fig_cancel_rate = px.bar(
            df_cancel_rate,
            x="fecha",
            y="canceladas",
            title="Número de Cancelaciones por Mes",
            labels={"fecha": "Mes", "canceladas": "Cancelaciones"},
            color="canceladas",
            color_continuous_scale="Reds"
        )
        st.plotly_chart(fig_cancel_rate, use_container_width=True)
    # --- Fin análisis cancelación ---

    # Gráficos principales
    st.subheader("📈 Tendencias Principales")
    
    # Obtener datos para gráficos
    reservas_por_dia, distribucion_canales, tipos_habitacion = get_datos_reservas(
        fecha_inicio, fecha_fin, canal_reserva
    )
    
    # Gráfico de reservas por día
    df_reservas = pd.DataFrame(reservas_por_dia)
    if not df_reservas.empty:
        fig_reservas = px.line(
            df_reservas,
            x="_id",
            y="total",
            title="Reservas por Día",
            labels={"_id": "Fecha", "total": "Número de Reservas"}
        )
        st.plotly_chart(fig_reservas, use_container_width=True)
    else:
        st.warning("No hay datos de reservas para el período seleccionado")
    
    # Gráficos en dos columnas
    col1, col2 = st.columns(2)
    
    with col1:
        # Gráfico de distribución por canal
        df_canales = pd.DataFrame(distribucion_canales)
        if not df_canales.empty:
            fig_canales = px.pie(
                df_canales,
                values="total",
                names="_id",
                title="Distribución por Canal de Reserva"
            )
            st.plotly_chart(fig_canales, use_container_width=True)
        else:
            st.warning("No hay datos de distribución por canal")
    
    with col2:
        # Gráfico de tipos de habitación
        df_habitaciones = pd.DataFrame(tipos_habitacion)
        if not df_habitaciones.empty:
            fig_habitaciones = px.bar(
                df_habitaciones,
                x="_id",
                y="total",
                title="Reservas por Tipo de Habitación",
                labels={"_id": "Tipo de Habitación", "total": "Número de Reservas"}
            )
            st.plotly_chart(fig_habitaciones, use_container_width=True)
        else:
            st.warning("No hay datos de tipos de habitación")

# Pestaña 2: Análisis Detallado
with tab2:
    st.header("📈 Análisis Detallado")
    
    # Obtener datos detallados
    datos_detallados = get_datos_detallados(fecha_inicio, fecha_fin, canal_reserva)
    
    if datos_detallados:
        df_detallados = pd.DataFrame(datos_detallados)
        df_detallados['fecha_llegada'] = pd.to_datetime(df_detallados['fecha_llegada'])
        df_detallados = df_detallados.sort_values('fecha_llegada', ascending=False)
        
        # Mostrar tabla de datos
        st.subheader("📋 Datos Detallados de Reservas")
        st.dataframe(df_detallados, use_container_width=True)
        
        # Análisis adicionales
        st.subheader("📊 Análisis Adicionales")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Análisis por país de origen
            if 'pais_origen' in df_detallados.columns:
                paises_count = df_detallados['pais_origen'].value_counts().head(10)
                fig_paises = px.bar(
                    x=paises_count.values,
                    y=paises_count.index,
                    orientation='h',
                    title="Top 10 Países de Origen",
                    labels={'x': 'Número de Reservas', 'y': 'País'}
                )
                st.plotly_chart(fig_paises, use_container_width=True)
        
        with col2:
            # Análisis de tiempo de anticipación
            if 'tiempo_anticipacion_reserva_dias' in df_detallados.columns:
                fig_anticipacion = px.histogram(
                    df_detallados,
                    x='tiempo_anticipacion_reserva_dias',
                    title="Distribución de Tiempo de Anticipación",
                    labels={'tiempo_anticipacion_reserva_dias': 'Días de Anticipación', 'count': 'Frecuencia'}
                )
                st.plotly_chart(fig_anticipacion, use_container_width=True)
        
        # Análisis de ADR por tipo de habitación
        if 'tipo_habitacion' in df_detallados.columns and 'adr' in df_detallados.columns:
            adr_por_tipo = df_detallados.groupby('tipo_habitacion')['adr'].mean().sort_values(ascending=False)
            fig_adr = px.bar(
                x=adr_por_tipo.index,
                y=adr_por_tipo.values,
                title="ADR Promedio por Tipo de Habitación",
                labels={'x': 'Tipo de Habitación', 'y': 'ADR Promedio ($)'}
            )
            st.plotly_chart(fig_adr, use_container_width=True)
    
    else:
        st.warning("No hay datos detallados para el período seleccionado")

# Pestaña 3: Análisis de Clientes
with tab3:
    st.header("👥 Análisis de Clientes")
    
    # Obtener análisis de clientes
    paises, recurrentes, top_clientes = get_analisis_clientes()
    
    # Clientes por país
    st.subheader("🌍 Distribución de Clientes por País")
    if paises:
        df_paises = pd.DataFrame(paises)
        fig_paises_clientes = px.pie(
            df_paises,
            values="total",
            names="_id",
            title="Distribución de Clientes por País"
        )
        st.plotly_chart(fig_paises_clientes, use_container_width=True)
    
    # Huéspedes recurrentes vs nuevos
    st.subheader("🔄 Huéspedes Recurrentes vs Nuevos")
    if recurrentes:
        df_recurrentes = pd.DataFrame(recurrentes)
        df_recurrentes['_id'] = df_recurrentes['_id'].map({True: 'Recurrentes', False: 'Nuevos'})
        fig_recurrentes = px.pie(
            df_recurrentes,
            values="total",
            names="_id",
            title="Proporción de Huéspedes Recurrentes"
        )
        st.plotly_chart(fig_recurrentes, use_container_width=True)


# Pestaña 4: Análisis de Ocupación
with tab4:
    st.header("🏠 Análisis de Ocupación")
    
    # Obtener análisis de ocupación
    datos_ocupacion = get_analisis_ocupacion(fecha_inicio, fecha_fin)
    
    if datos_ocupacion:
        df_ocupacion = pd.DataFrame(datos_ocupacion)
        df_ocupacion['fecha'] = pd.to_datetime(df_ocupacion['_id'].apply(lambda x: f"{x['anio']}-{x['mes']:02d}-01"))
        
        # Gráfico de ocupación por mes
        fig_ocupacion = px.line(
            df_ocupacion,
            x='fecha',
            y='total_reservas',
            title="Ocupación Mensual",
            labels={'fecha': 'Mes', 'total_reservas': 'Número de Reservas'}
        )
        st.plotly_chart(fig_ocupacion, use_container_width=True)
        
        # Gráfico de ADR por mes
        fig_adr_mensual = px.line(
            df_ocupacion,
            x='fecha',
            y='adr_promedio',
            title="ADR Promedio Mensual",
            labels={'fecha': 'Mes', 'adr_promedio': 'ADR Promedio ($)'}
        )
        st.plotly_chart(fig_adr_mensual, use_container_width=True)
        
        # Gráfico de noches totales
        fig_noches = px.bar(
            df_ocupacion,
            x='fecha',
            y='total_noches',
            title="Total de Noches por Mes",
            labels={'fecha': 'Mes', 'total_noches': 'Total de Noches'}
        )
        st.plotly_chart(fig_noches, use_container_width=True)
    
    else:
        st.warning("No hay datos de ocupación para el período seleccionado") 

# Pestaña 5: Nueva Reserva
with tab5:
    st.header("➕ Agregar Nueva Reserva")
    st.info("Completa el formulario para registrar una nueva reserva. Si el cliente ya existe por email o documento, se reutilizará.")

    with st.form("form_nueva_reserva"):
        st.subheader("Datos del Cliente")
        nombre_cliente = st.text_input("Nombre completo", max_chars=100)
        email_cliente = st.text_input("Email", max_chars=100)
        telefono_cliente = st.text_input("Teléfono", max_chars=30)
        tipo_doc = st.text_input("Tipo de documento", max_chars=30)
        num_doc = st.text_input("Número de documento", max_chars=30)
        fecha_nac = st.date_input("Fecha de nacimiento", value=None)
        pais_cliente = st.text_input("País de origen", max_chars=50)
        es_recurrente = st.checkbox("¿Es huésped recurrente?", value=False)

        st.subheader("Datos de la Reserva")
        fecha_llegada = st.date_input("Fecha de llegada")
        noches_estadia = st.number_input("Noches de estadía", min_value=1, value=1)
        fecha_salida = fecha_llegada + timedelta(days=noches_estadia)
        canal_reserva = st.selectbox("Canal de reserva", options=get_canales_reserva()[1:])
        adr = st.number_input("ADR (Tarifa Diaria Promedio)", min_value=0.0, value=100.0)
        estado_reserva = st.text_input("Estado de la reserva", value="Confirmada")
        fue_cancelada = st.checkbox("¿Fue cancelada?", value=False)
        tiempo_anticipacion = st.number_input("Días de anticipación de la reserva", min_value=0, value=0)
        fecha_creacion = st.date_input("Fecha de creación de la reserva", value=datetime.now().date())
        fecha_estado = st.date_input("Fecha de estado de la reserva", value=datetime.now().date())

        st.subheader("Detalles de la Reserva")
        tipo_habitacion = st.text_input("Tipo de habitación reservada", max_chars=30)
        tipo_habitacion_asignada = st.text_input("Tipo de habitación asignada", max_chars=30)
        cambios_en_reserva = st.number_input("Cambios en la reserva", min_value=0, value=0)
        tipo_cliente_en_reserva = st.text_input("Tipo de cliente en reserva", max_chars=30)

        submitted = st.form_submit_button("Guardar Reserva")

        if submitted:
            # Validación básica
            if not nombre_cliente or not email_cliente or not fecha_llegada or not tipo_habitacion:
                st.error("Por favor, completa los campos obligatorios marcados con *.")
                st.stop()

            # Buscar o crear cliente
            cliente = db.Clientes.find_one({"email": email_cliente})
            if not cliente and tipo_doc and num_doc:
                cliente = db.Clientes.find_one({
                    "tipo_documento_identidad": tipo_doc,
                    "numero_documento_identidad": num_doc
                })
            if not cliente:
                cliente_id = ObjectId()
                cliente_data = {
                    "_id": cliente_id,
                    "nombre_completo": nombre_cliente,
                    "email": email_cliente,
                    "telefono": telefono_cliente,
                    "tipo_documento_identidad": tipo_doc,
                    "numero_documento_identidad": num_doc,
                    "fecha_nacimiento": datetime.combine(fecha_nac, datetime.min.time()) if fecha_nac else None,
                    "pais_origen_cliente": pais_cliente,
                    "es_huesped_recurrente_historico": es_recurrente,
                    "total_cancelaciones_previas_cliente": 0,
                    "total_reservas_previas_no_canceladas_cliente": 0,
                    "historial_ids_reservas": []
                }
                db.Clientes.insert_one(cliente_data)
                st.success(f"Cliente creado: {nombre_cliente}")
            else:
                cliente_id = cliente["_id"]
                st.info(f"Cliente existente reutilizado: {cliente['nombre_completo']}")

            # Crear detalle de reserva
            detalle_id = ObjectId()
            detalle_data = {
                "_id": detalle_id,
                "reserva_id": None,  # Se asigna después
                "pais_origen_reserva": pais_cliente,
                "es_huesped_recurrente_al_reservar": es_recurrente,
                "cancelaciones_previas_cliente_al_reservar": 0,
                "reservas_previas_no_canceladas_cliente_al_reservar": 0,
                "tipo_habitacion_reservada": tipo_habitacion,
                "tipo_habitacion_asignada": tipo_habitacion_asignada,
                "cambios_en_reserva": cambios_en_reserva,
                "tipo_cliente_en_reserva": tipo_cliente_en_reserva
            }
            db.DetallesReserva.insert_one(detalle_data)

            # Crear reserva
            reserva_id = ObjectId()
            db.DetallesReserva.update_one({"_id": detalle_id}, {"$set": {"reserva_id": reserva_id}})
            reserva_data = {
                "_id": reserva_id,
                "cliente_id": cliente_id,
                "detalle_reserva_id": detalle_id,
                "fecha_creacion_reserva": datetime.combine(fecha_creacion, datetime.min.time()),
                "fue_cancelada": fue_cancelada,
                "tiempo_anticipacion_reserva_dias": tiempo_anticipacion,
                "fecha_llegada": datetime.combine(fecha_llegada, datetime.min.time()),
                "fecha_salida": datetime.combine(fecha_salida, datetime.min.time()),
                "noches_estadia": noches_estadia,
                "estado_reserva": estado_reserva,
                "fecha_estado_reserva": datetime.combine(fecha_estado, datetime.min.time()),
                "adr": adr,
                "canal_reserva": canal_reserva
            }
            db.Reservas.insert_one(reserva_data)
            db.Clientes.update_one({"_id": cliente_id}, {"$push": {"historial_ids_reservas": reserva_id}})

            st.success("¡Reserva guardada exitosamente!")
            st.balloons() 

# Pestaña 6: Búsqueda y Gestión de Reservas
with tab6:
    st.header("🔍 Búsqueda y Gestión de Reservas")
    
    # Filtros de búsqueda
    col1, col2, col3 = st.columns(3)
    
    with col1:
        busqueda_cliente = st.text_input("Buscar por nombre o email del cliente")
    
    with col2:
        estado_reserva = st.selectbox(
            "Estado de la reserva",
            options=["Todos", "Confirmada", "Cancelada", "Pendiente", "Completada"]
        )
    
    with col3:
        tipo_habitacion_filtro = st.selectbox(
            "Tipo de habitación",
            options=["Todos"] + get_tipos_habitacion()[1:]
        )
    
    # Botón de búsqueda
    if st.button("🔍 Buscar Reservas"):
        # Construir filtro de búsqueda
        filtro_busqueda = {}
        
        if busqueda_cliente:
            # Buscar cliente por nombre o email
            pipeline_cliente = [
                {"$match": {
                    "$or": [
                        {"nombre_completo": {"$regex": busqueda_cliente, "$options": "i"}},
                        {"email": {"$regex": busqueda_cliente, "$options": "i"}}
                    ]
                }},
                {"$project": {"_id": 1}}
            ]
            clientes_encontrados = list(db.Clientes.aggregate(pipeline_cliente))
            if clientes_encontrados:
                cliente_ids = [c["_id"] for c in clientes_encontrados]
                filtro_busqueda["cliente_id"] = {"$in": cliente_ids}
        
        if estado_reserva != "Todos":
            filtro_busqueda["estado_reserva"] = estado_reserva
        
        # Pipeline para obtener reservas con detalles
        pipeline_reservas = [
            {"$match": filtro_busqueda},
            {
                "$lookup": {
                    "from": "DetallesReserva",
                    "localField": "detalle_reserva_id",
                    "foreignField": "_id",
                    "as": "detalle"
                }
            },
            {"$unwind": "$detalle"},
            {
                "$lookup": {
                    "from": "Clientes",
                    "localField": "cliente_id",
                    "foreignField": "_id",
                    "as": "cliente"
                }
            },
            {"$unwind": "$cliente"},
            {
                "$project": {
                    "fecha_llegada": 1,
                    "fecha_salida": 1,
                    "noches_estadia": 1,
                    "adr": 1,
                    "canal_reserva": 1,
                    "fue_cancelada": 1,
                    "estado_reserva": 1,
                    "nombre_cliente": "$cliente.nombre_completo",
                    "email_cliente": "$cliente.email",
                    "tipo_habitacion": "$detalle.tipo_habitacion_reservada",
                    "tipo_cliente": "$detalle.tipo_cliente_en_reserva"
                }
            },
            {"$sort": {"fecha_llegada": -1}},
            {"$limit": 100}
        ]
        
        reservas_encontradas = list(db.Reservas.aggregate(pipeline_reservas))
        
        if reservas_encontradas:
            df_reservas = pd.DataFrame(reservas_encontradas)
            df_reservas['fecha_llegada'] = pd.to_datetime(df_reservas['fecha_llegada'])
            df_reservas['fecha_salida'] = pd.to_datetime(df_reservas['fecha_salida'])
            
            # Filtrar por tipo de habitación si se especifica
            if tipo_habitacion_filtro != "Todos":
                df_reservas = df_reservas[df_reservas['tipo_habitacion'] == tipo_habitacion_filtro]
            
            st.success(f"✅ Se encontraron {len(df_reservas)} reservas")
            
            # Mostrar tabla de resultados
            st.subheader("📋 Resultados de la Búsqueda")
            st.dataframe(df_reservas, use_container_width=True)
            
            # Estadísticas de la búsqueda
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Reservas", len(df_reservas))
            with col2:
                canceladas = df_reservas['fue_cancelada'].sum()
                st.metric("Canceladas", canceladas)
            with col3:
                adr_promedio = df_reservas['adr'].mean()
                st.metric("ADR Promedio", f"${adr_promedio:.2f}")
            with col4:
                noches_promedio = df_reservas['noches_estadia'].mean()
                st.metric("Noches Promedio", f"{noches_promedio:.1f}")
            
        else:
            st.warning("No se encontraron reservas con los criterios especificados")

# Pestaña 7: Reportes y Exportación
with tab7:
    st.header("📈 Reportes y Exportación")
    
    # Selección de tipo de reporte
    tipo_reporte = st.selectbox(
        "Seleccionar tipo de reporte",
        options=[
            "📊 Reporte General de Reservas",
            "💰 Reporte de Rentabilidad",
            "👥 Reporte de Clientes",
            "❌ Reporte de Cancelaciones",
            "📅 Reporte Mensual"
        ]
    )
    
    # Configuración del reporte
    col1, col2 = st.columns(2)
    
    with col1:
        fecha_inicio_reporte = st.date_input(
            "Fecha inicio reporte",
            value=fecha_inicio.date()
        )
    
    with col2:
        fecha_fin_reporte = st.date_input(
            "Fecha fin reporte",
            value=fecha_fin.date()
        )
    
    # Generar reporte
    if st.button("📊 Generar Reporte"):
        fecha_inicio_dt = datetime.combine(fecha_inicio_reporte, datetime.min.time())
        fecha_fin_dt = datetime.combine(fecha_fin_reporte, datetime.max.time())
        
        if tipo_reporte == "📊 Reporte General de Reservas":
            # Pipeline para reporte general
            pipeline_general = [
                {"$match": {
                    "fecha_llegada": {"$gte": fecha_inicio_dt, "$lte": fecha_fin_dt}
                }},
                {
                    "$lookup": {
                        "from": "DetallesReserva",
                        "localField": "detalle_reserva_id",
                        "foreignField": "_id",
                        "as": "detalle"
                    }
                },
                {"$unwind": "$detalle"},
                {
                    "$lookup": {
                        "from": "Clientes",
                        "localField": "cliente_id",
                        "foreignField": "_id",
                        "as": "cliente"
                    }
                },
                {"$unwind": "$cliente"},
                {
                    "$project": {
                        "fecha_llegada": 1,
                        "fecha_salida": 1,
                        "noches_estadia": 1,
                        "adr": 1,
                        "canal_reserva": 1,
                        "fue_cancelada": 1,
                        "estado_reserva": 1,
                        "nombre_cliente": "$cliente.nombre_completo",
                        "email_cliente": "$cliente.email",
                        "tipo_habitacion": "$detalle.tipo_habitacion_reservada",
                        "tipo_cliente": "$detalle.tipo_cliente_en_reserva"
                    }
                }
            ]
            
            datos_reporte = list(db.Reservas.aggregate(pipeline_general))
            
            if datos_reporte:
                df_reporte = pd.DataFrame(datos_reporte)
                df_reporte['fecha_llegada'] = pd.to_datetime(df_reporte['fecha_llegada'])
                df_reporte['fecha_salida'] = pd.to_datetime(df_reporte['fecha_salida'])
                
                st.success(f"✅ Reporte generado con {len(df_reporte)} registros")
                
                # Métricas del reporte
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Reservas", len(df_reporte))
                with col2:
                    canceladas = df_reporte['fue_cancelada'].sum()
                    st.metric("Canceladas", canceladas)
                with col3:
                    adr_promedio = df_reporte['adr'].mean()
                    st.metric("ADR Promedio", f"${adr_promedio:.2f}")
                with col4:
                    ingresos_totales = (df_reporte['adr'] * df_reporte['noches_estadia']).sum()
                    st.metric("Ingresos Totales", f"${ingresos_totales:,.2f}")
                
                # Mostrar datos
                st.subheader("📋 Datos del Reporte")
                st.dataframe(df_reporte, use_container_width=True)
                
                # Botones de exportación
                col1, col2 = st.columns(2)
                
                with col1:
                    # Exportar CSV
                    csv = df_reporte.to_csv(index=False)
                    st.download_button(
                        label="📥 Descargar CSV",
                        data=csv,
                        file_name=f"reporte_general_{fecha_inicio_reporte}_{fecha_fin_reporte}.csv",
                        mime="text/csv"
                    )
                
                with col2:
                    # Exportar Excel
                    def export_to_excel():
                        # Limpiar datos para Excel
                        df_clean = clean_data_for_excel(df_reporte)
                        
                        # Crear workbook
                        wb = Workbook()
                        
                        # Hoja 1: Datos principales
                        ws1 = wb.active
                        ws1.title = "Datos de Reservas"
                        
                        # Agregar título
                        ws1['A1'] = f"Reporte General de Reservas - {fecha_inicio_reporte} a {fecha_fin_reporte}"
                        ws1['A1'].font = Font(bold=True, size=14)
                        ws1.merge_cells('A1:K1')
                        
                        # Agregar métricas
                        ws1['A3'] = "Métricas del Reporte"
                        ws1['A3'].font = Font(bold=True, size=12)
                        
                        ws1['A4'] = "Total Reservas"
                        ws1['B4'] = len(df_clean)
                        ws1['A5'] = "Reservas Canceladas"
                        ws1['B5'] = canceladas
                        ws1['A6'] = "ADR Promedio"
                        ws1['B6'] = adr_promedio
                        ws1['A7'] = "Ingresos Totales"
                        ws1['B7'] = ingresos_totales
                        
                        # Agregar datos
                        ws1['A9'] = "Datos Detallados"
                        ws1['A9'].font = Font(bold=True, size=12)
                        
                        # Headers
                        headers = list(df_clean.columns)
                        for col, header in enumerate(headers, 1):
                            cell = ws1.cell(row=10, column=col)
                            cell.value = header
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        # Datos
                        for row_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=False), 11):
                            for col_idx, value in enumerate(row, 1):
                                ws1.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Ajustar ancho de columnas de forma segura
                        for col_num in range(1, len(headers) + 1):
                            max_length = 0
                            # Usar la fila 10 (donde están los headers) para obtener la letra de columna
                            column_letter = ws1.cell(row=10, column=col_num).column_letter
                            
                            for row_num in range(1, ws1.max_row + 1):
                                cell = ws1.cell(row=row_num, column=col_num)
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            ws1.column_dimensions[column_letter].width = adjusted_width
                        
                        # Hoja 2: Resumen por canal
                        ws2 = wb.create_sheet("Resumen por Canal")
                        canal_summary = df_reporte['canal_reserva'].value_counts()
                        
                        ws2['A1'] = "Resumen por Canal de Reserva"
                        ws2['A1'].font = Font(bold=True, size=14)
                        
                        ws2['A3'] = "Canal"
                        ws2['B3'] = "Cantidad"
                        ws2['A3'].font = Font(bold=True)
                        ws2['B3'].font = Font(bold=True)
                        
                        for idx, (canal, cantidad) in enumerate(canal_summary.items(), 4):
                            ws2[f'A{idx}'] = canal
                            ws2[f'B{idx}'] = cantidad
                        
                        # Hoja 3: Resumen por tipo de habitación
                        ws3 = wb.create_sheet("Resumen por Habitación")
                        habitacion_summary = df_reporte['tipo_habitacion'].value_counts()
                        
                        ws3['A1'] = "Resumen por Tipo de Habitación"
                        ws3['A1'].font = Font(bold=True, size=14)
                        
                        ws3['A3'] = "Tipo de Habitación"
                        ws3['B3'] = "Cantidad"
                        ws3['A3'].font = Font(bold=True)
                        ws3['B3'].font = Font(bold=True)
                        
                        for idx, (habitacion, cantidad) in enumerate(habitacion_summary.items(), 4):
                            ws3[f'A{idx}'] = habitacion
                            ws3[f'B{idx}'] = cantidad
                        
                        # Guardar en buffer
                        buffer = io.BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        return buffer.getvalue()
                    
                    excel_data = export_to_excel()
                    st.download_button(
                        label="📊 Descargar Excel",
                        data=excel_data,
                        file_name=f"reporte_general_{fecha_inicio_reporte}_{fecha_fin_reporte}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                # Gráfico resumen
                st.subheader("📊 Resumen Visual")
                col1, col2 = st.columns(2)
                
                with col1:
                    # Distribución por canal
                    canal_counts = df_reporte['canal_reserva'].value_counts()
                    fig_canal = px.pie(
                        values=canal_counts.values,
                        names=canal_counts.index,
                        title="Distribución por Canal"
                    )
                    st.plotly_chart(fig_canal, use_container_width=True)
                
                with col2:
                    # Distribución por tipo de habitación
                    habitacion_counts = df_reporte['tipo_habitacion'].value_counts()
                    fig_habitacion = px.bar(
                        x=habitacion_counts.index,
                        y=habitacion_counts.values,
                        title="Distribución por Tipo de Habitación"
                    )
                    st.plotly_chart(fig_habitacion, use_container_width=True)
            
            else:
                st.warning("No hay datos para el período seleccionado")
        
        elif tipo_reporte == "💰 Reporte de Rentabilidad":
            # Pipeline para reporte de rentabilidad
            pipeline_rentabilidad = [
                {"$match": {
                    "fecha_llegada": {"$gte": fecha_inicio_dt, "$lte": fecha_fin_dt}
                }},
                {
                    "$lookup": {
                        "from": "DetallesReserva",
                        "localField": "detalle_reserva_id",
                        "foreignField": "_id",
                        "as": "detalle"
                    }
                },
                {"$unwind": "$detalle"},
                {
                    "$group": {
                        "_id": {
                            "tipo_habitacion": "$detalle.tipo_habitacion_reservada",
                            "canal_reserva": "$canal_reserva"
                        },
                        "total_reservas": {"$sum": 1},
                        "total_noches": {"$sum": "$noches_estadia"},
                        "adr_promedio": {"$avg": "$adr"},
                        "ingresos_totales": {"$sum": {"$multiply": ["$adr", "$noches_estadia"]}},
                        "cancelaciones": {"$sum": {"$cond": ["$fue_cancelada", 1, 0]}}
                    }
                },
                {"$sort": {"ingresos_totales": -1}}
            ]
            
            datos_rentabilidad = list(db.Reservas.aggregate(pipeline_rentabilidad))
            
            if datos_rentabilidad:
                df_rentabilidad = pd.DataFrame(datos_rentabilidad)
                df_rentabilidad['tasa_cancelacion'] = (df_rentabilidad['cancelaciones'] / df_rentabilidad['total_reservas'] * 100)
                
                st.success(f"✅ Reporte de rentabilidad generado")
                
                # Mostrar tabla de rentabilidad
                st.subheader("💰 Análisis de Rentabilidad")
                st.dataframe(df_rentabilidad, use_container_width=True)
                
                # Gráfico de ingresos por tipo de habitación
                ingresos_por_tipo = df_rentabilidad.groupby('_id')['ingresos_totales'].sum().sort_values(ascending=False)
                fig_ingresos = px.bar(
                    x=ingresos_por_tipo.index,
                    y=ingresos_por_tipo.values,
                    title="Ingresos por Tipo de Habitación y Canal"
                )
                st.plotly_chart(fig_ingresos, use_container_width=True)
                
                # Botones de exportación
                col1, col2 = st.columns(2)
                
                with col1:
                    # Exportar CSV
                    csv = df_rentabilidad.to_csv(index=False)
                    st.download_button(
                        label="📥 Descargar CSV",
                        data=csv,
                        file_name=f"reporte_rentabilidad_{fecha_inicio_reporte}_{fecha_fin_reporte}.csv",
                        mime="text/csv"
                    )
                
                with col2:
                    # Exportar Excel
                    def export_rentabilidad_to_excel():
                        # Limpiar datos para Excel
                        df_clean = clean_data_for_excel(df_rentabilidad)
                        
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Reporte de Rentabilidad"
                        
                        # Título
                        ws['A1'] = f"Reporte de Rentabilidad - {fecha_inicio_reporte} a {fecha_fin_reporte}"
                        ws['A1'].font = Font(bold=True, size=14)
                        ws.merge_cells('A1:F1')
                        
                        # Headers
                        headers = list(df_clean.columns)
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=3, column=col)
                            cell.value = header
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        # Datos
                        for row_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=False), 4):
                            for col_idx, value in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Ajustar columnas de forma segura
                        for col_num in range(1, len(headers) + 1):
                            max_length = 0
                            # Usar la fila 3 (donde están los headers) para obtener la letra de columna
                            column_letter = ws.cell(row=3, column=col_num).column_letter
                            
                            for row_num in range(1, ws.max_row + 1):
                                cell = ws.cell(row=row_num, column=col_num)
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        buffer = io.BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)
                        return buffer.getvalue()
                    
                    excel_data = export_rentabilidad_to_excel()
                    st.download_button(
                        label="📊 Descargar Excel",
                        data=excel_data,
                        file_name=f"reporte_rentabilidad_{fecha_inicio_reporte}_{fecha_fin_reporte}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            else:
                st.warning("No hay datos de rentabilidad para el período seleccionado")
        
        # Agregar más tipos de reporte según sea necesario...
        else:
            st.info("Este tipo de reporte estará disponible próximamente") 